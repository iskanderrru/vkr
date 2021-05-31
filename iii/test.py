'''
import openpyxl


data = open('data.ini', 'w')
formuls = open('formulas.ini', 'w')
data.write('[DATA]\n')
formuls.write('[FORMULS]\n')
xl_file = openpyxl.load_workbook("./Transform.xlsx")
monitor = xl_file.active
for i in range(1, 246):
	if monitor[f'D{i}'].value:
		if monitor[f'D{i}'].value.startswith('='):
			data.write(f'D{i}' + str(monitor[f'D{i}'].value).replace('%', 'p') + '\n')

for i in range(1, 246):
	if monitor[f'E{i}'].value:
		if monitor[f'D{i}'].value.startswith('='):
			formuls.write(f'curr{i}' + str(monitor[f'E{i}'].value).replace('%', 'p').replace('D', 'prev').replace('E', 'curr') + '\n')
'''
import formulas
import openpyxl
import configparser


years = {
			0: 'D',
			1: 'E',
			2: 'F',
			3: 'G',
			4: 'H',
			5: 'I',
			6: 'J',
			7: 'K',
			8: 'L',
			9: 'M',
			10: 'N',
			11: 'O',
			12: 'P',
			13: 'Q',
			14: 'R',
			15: 'S',
			16: 'T',
			17: 'U',
			18: 'V',
			19: 'W',
			20: 'X',
			21: 'Y',
			22: 'Z',
			23: 'AA',
			24: 'AB',
			25: 'AC',
		}
curryear = 1
formuls = configparser.ConfigParser()
formuls.read('formulas.ini')
xlfile = openpyxl.load_workbook('base.xlsx')
ws = xlfile.active

def prepformula(formula):
	global years, curryear
	formula = formula.replace('curr', years[curryear]).replace('prev', years[curryear - 1])
	if formula.endswith('p'):
		formula = float(formula[:-1]) / 100

	return '=' + str(formula)

def prepdata(value):
	value = value.replace('=', '') if str(value).startswith('=') else value
	if str(value).endswith('%'):
		return float(value[:-1]) / 100
	else:
		return float(value)

def myfunc(cell):
	global formuls, years, curryear, ws
	args = dict()
	formula = prepformula(formuls['FORMULS'][cell.replace(years[curryear], 'curr')])
	func = formulas.Parser().ast(formula)[1].compile()
	for inp in list(func.inputs):
		if inp.startswith(years[curryear]):
			args[inp] = myfunc(inp)
		else:
			dat = prepdata(ws[inp.replace(years[curryear], years[curryear - 1])].value)
			args[inp] = dat
	#print(cell, args)
	return func(**args)

def main(n):
	global years, curryear, formuls, ws
	for i in range(1, n):
		print(f'Computing year {i}')
		for item in formuls['FORMULS']:
			val = float(myfunc(item))
			val = float(f'{val:.4f}')
			ws[f'{years[curryear]}{item[4:]}'] = val
		curryear += 1
		print(f'Year {i} done')

	xlfile.save('res.xlsx')
	print('Results saved to res.xlsx')

if __name__ == '__main__':
	main(int(input('Input amount of years to compute: ')) + 1)

